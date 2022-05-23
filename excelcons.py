import shutil
import os.path
from os import path
import xlsxwriter
import openpyxl as xl
#from openpyxl import Workbook, load_workbook
#from openpyxl.styles import Border, Side, PatternFill
from copy import copy

def copyfiles(monitor, destin):
    #print("copy: "+monitor) -file to be copied
    #print(monitor.endswith('master.xlsx')) -result to copy or not file
    if monitor.endswith('master.xlsx') == False:
        shutil.move(monitor, destin)

def fileprocess(ffile):
    #print(ffile)
    master_is = os.path.join(monitor, "master.xlsx")
    if os.path.exists(master_is) == True:
        print("Master file found")
    else:
        print("Master file will be created")
        workbook = xlsxwriter.Workbook(master_is)
        worksheet = workbook.add_worksheet()
        workbook.close()
    
    original_wb = load_workbook(ffile)
    copy_to_wb = load_workbook(master_is)
    nosheet = 0
    for isheet in original_wb.worksheets:
        
        source_sheet = original_wb.worksheets[nosheet]
        copy_to_sheet = copy_to_wb.create_sheet(source_sheet.title+"_copy")
        nosheet = nosheet + 1
        for row in source_sheet:
            for cell in row:
                copy_to_sheet[cell.coordinate].value = cell.value
                copy_to_sheet[cell.coordinate].style = copy(cell.style)
                copy_to_sheet[cell.coordinate].border = copy(cell.border)
                copy_to_sheet[cell.coordinate].fill = copy(cell.fill)
                copy_to_sheet[cell.coordinate].alignment = copy(cell.alignment)
                
                
    copy_to_wb.save(master_is)


i = 0
#receiving and verifying input path for folder to monitor
while i!=1:
    print('Path of folder to monitor: ')
    monitor = input() #general varaible of folder to monitor
    print('Is the path correct? y/n: ' + monitor)
    confirmation = input()
    if confirmation == 'Y' or confirmation == 'y':
        if path.isdir(monitor) == True:
            i = 1
            print('Path accepted')
        else:
            i = 0
            print('path not accepted\n')
    else:
        i = 0
        print('path not accepted\n')



#defining and creating paths for folders
if i==1:
    dirnotapplicable = "Not Applicable"
    dirprocessed = "Processed"
    fpathapplicable = os.path.join(monitor, dirnotapplicable)
    fpathprocessed = os.path.join(monitor, dirprocessed)

    if path.exists(fpathapplicable) == False and path.exists(fpathprocessed) == False:
      os.mkdir(fpathapplicable)
      os.mkdir(fpathprocessed)

 #start of program to keep processing files

#begin monitoring
before = dict ([(f, None) for f in os.listdir (monitor)])
while 1: 
        after = dict ([(f, None) for f in os.listdir (monitor)])
        added = [f for f in after if not f in before]
        if added:
                print("Added: ", ", ".join (added))
                before = after
                #start to check if file is xls or xlsx
                filetomove = os.path.join(monitor, added[0])
                #print(filetomove, " xxx ", added[0])
                ext = os.path.splitext(added[0])[-1].lower()
                if ext == ".xls" or ext== ".xlsx":
                    is_excel = True
                else:
                    is_excel = False
                  
                
                if is_excel == True:
                    print("Processing file" + filetomove)
                    fileprocess(filetomove)
                    copyfiles(filetomove, fpathprocessed)
                else:
                    print("File not applicable")
                    copyfiles(filetomove, fpathapplicable)
        else:
             before = after       
    #os.rmdir(fpathapplicable) #debug
    #os.rmdir(fpathprocessed)



